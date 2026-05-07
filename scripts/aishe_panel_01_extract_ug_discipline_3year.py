"""
Build a 3-year AISHE panel of UG enrolment + UG out-turn by major discipline.

Sources (all in sources/):
  aishe_2019-20_final_report.xlsx
  aishe_2020-21_final_report.xlsx
  aishe_2021-22_final_report.xlsx

For each AISHE Final Report Excel, we read:
  - Sheet '12UGDisc ' (note trailing space) -> UG ENROLMENT by Major Discipline
  - Sheet '35UGDisc'                        -> UG OUT-TURN by Major Discipline

Both sheets have similar layouts but column positions differ slightly across
years (S.No. column was added in 2021-22). We auto-detect.

Output:
  extractions/aishe_ug_discipline_panel_2019-22.csv
Schema: aishe_year, metric (enrolment | out_turn), discipline, gender, value
"""
import csv
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SOURCES = ROOT / "sources"
OUT = ROOT / "extractions" / "aishe_ug_discipline_panel_2019-22.csv"

REPORTS = [
    ("2019-20", SOURCES / "aishe_2019-20_final_report.xlsx"),
    ("2020-21", SOURCES / "aishe_2020-21_final_report.xlsx"),
    ("2021-22", SOURCES / "aishe_2021-22_final_report.xlsx"),
]


def extract_discipline_table(ws, metric: str, year: str):
    """Yield rows for a UG-by-discipline sheet.

    Layout varies: 2019-20 / 2020-21 had no S.No. column (Discipline at col 1,
    Subject at col 2, Male/Female/Total at cols 3/4/5). 2021-22 has S.No. at
    col 1 (Discipline at col 2, Subject at col 3, Male/Female/Total at cols 4/5/6).

    A discipline's TOTAL row has Subject column empty (or contains the
    discipline name with 'Total' suffix). We keep only those rows.
    """
    rows = []
    # Auto-detect schema by finding the row whose first or second cell EQUALS
    # 'Discipline' exactly (not the title row which has "Disciplines/ Subjects").
    # The header row also has 'Male' / 'Female' / 'Total' as nearby cells.
    schema = None
    for ri in range(1, 6):
        cells = [str(c.value).strip() if c.value is not None else "" for c in ws[ri]]
        if not cells:
            continue
        # Old schema: Discipline at col 0 (index 0)
        if cells[0] == "Discipline":
            schema = "old"
            break
        # New schema: S.No. at col 0, Discipline at col 1
        if len(cells) >= 2 and cells[1] == "Discipline":
            schema = "new"
            break
    if schema is None:
        raise RuntimeError(f"could not detect schema in sheet {ws.title!r}")

    if schema == "old":
        col_disc, col_subj, col_m, col_f, col_t = 0, 1, 2, 3, 4
        first_data_row = 4
    else:
        col_disc, col_subj, col_m, col_f, col_t = 1, 2, 3, 4, 5
        first_data_row = 4

    last_disc = None
    for r in ws.iter_rows(min_row=first_data_row, values_only=True):
        if not r or len(r) <= col_t:
            continue
        disc = r[col_disc]
        subj = r[col_subj]
        male = r[col_m]
        female = r[col_f]
        total = r[col_t]
        if disc is None or not str(disc).strip():
            continue
        disc_s = str(disc).strip()
        # Some sheets have S.No row [1,2,3,4,5] interleaved — skip
        if str(disc_s).isdigit():
            continue
        # Discipline-total rows: subject column empty (None or "")
        # OR discipline name ends with "Total"
        is_total = (subj is None or str(subj).strip() == "")
        is_total = is_total or disc_s.endswith("Total")
        if not is_total:
            # Sub-discipline row — skip
            continue
        # Strip "Total" suffix if present
        clean = disc_s
        if clean.endswith("Total"):
            clean = clean[:-len("Total")].strip()
        # Skip blanks / row-number rows
        if not clean:
            continue
        if not isinstance(total, (int, float)):
            continue
        rows.append({
            "aishe_year": year,
            "metric": metric,
            "discipline": clean,
            "gender": "Male",
            "value": int(male) if isinstance(male, (int, float)) else 0,
        })
        rows.append({
            "aishe_year": year,
            "metric": metric,
            "discipline": clean,
            "gender": "Female",
            "value": int(female) if isinstance(female, (int, float)) else 0,
        })
        rows.append({
            "aishe_year": year,
            "metric": metric,
            "discipline": clean,
            "gender": "Total",
            "value": int(total) if isinstance(total, (int, float)) else 0,
        })
    return rows


def main():
    all_rows = []
    for year, pdf in REPORTS:
        wb = openpyxl.load_workbook(pdf, data_only=True)
        # Find the right sheet names — they vary in trailing space
        ug_enrol_sheet = None
        ug_outturn_sheet = None
        for s in wb.sheetnames:
            sl = s.replace(" ", "").lower()
            if sl == "12ugdisc":
                ug_enrol_sheet = s
            if sl == "35ugdisc":
                ug_outturn_sheet = s
        if not (ug_enrol_sheet and ug_outturn_sheet):
            print(f"WARN {year}: missing sheet (enrol={ug_enrol_sheet!r}, outturn={ug_outturn_sheet!r})")
            continue
        enrol_rows = extract_discipline_table(wb[ug_enrol_sheet], "enrolment", year)
        outturn_rows = extract_discipline_table(wb[ug_outturn_sheet], "out_turn", year)
        all_rows.extend(enrol_rows + outturn_rows)
        print(
            f"  {year}: {len({(r['discipline']) for r in enrol_rows})} disciplines (enrolment), "
            f"{len({(r['discipline']) for r in outturn_rows})} disciplines (out-turn)"
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["aishe_year", "metric", "discipline", "gender", "value"])
        w.writeheader()
        w.writerows(all_rows)
    print(f"\nWrote {len(all_rows):,} rows -> {OUT}")

    # Quick sanity check: print discipline totals across years
    from collections import defaultdict
    by_yr_disc = defaultdict(int)
    for r in all_rows:
        if r["metric"] == "out_turn" and r["gender"] == "Total":
            by_yr_disc[(r["aishe_year"], r["discipline"])] = r["value"]
    disciplines = sorted({d for _, d in by_yr_disc})
    print(f"\nSanity: UG out-turn (Total) for top disciplines, year-by-year:")
    print(f"{'Discipline':30s} {'2019-20':>12s} {'2020-21':>12s} {'2021-22':>12s}  CAGR")
    for d in disciplines:
        v19 = by_yr_disc.get(("2019-20", d), 0)
        v20 = by_yr_disc.get(("2020-21", d), 0)
        v21 = by_yr_disc.get(("2021-22", d), 0)
        if max(v19, v20, v21) < 50000:
            continue
        if v19 > 0 and v21 > 0:
            cagr = ((v21 / v19) ** (1 / 2) - 1) * 100
            cagr_s = f"{cagr:+5.1f}%"
        else:
            cagr_s = "—"
        print(f"{d:30s} {v19:>12,} {v20:>12,} {v21:>12,}  {cagr_s}")


if __name__ == "__main__":
    main()
